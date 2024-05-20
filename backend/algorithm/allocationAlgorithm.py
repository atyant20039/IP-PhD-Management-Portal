import re
import os
import math
import openpyxl
import argparse
import pandas as pd
from course import Course
from datetime import datetime
from django.conf import settings
from phdStudent import PhDStudent

#########################################################################################################################
################################################ Command Line Arguements ################################################
#########################################################################################################################

# Create an ArgumentParser object
parser = argparse.ArgumentParser()

# Add arguments
parser.add_argument('--TARatio', type=int, help='Value of TARatio')

# Parse the command-line arguments
args = parser.parse_args()

# Access the value of TARatio
TARatio = args.TARatio

#########################################################################################################################
################################################### Utility Functions ###################################################
#########################################################################################################################

def getExamDate(course_pool, course_code):
    """
    Returns the exam date for a specified course.
    
    This function takes a pool of courses and a specific course code, and returns 
    the exam date for the course corresponding to the given course code. It 
    retrieves the course object from the course pool using the course code as the key, 
    and then calls the `get_date()` method on the course object to get the exam date.
    
    Args:
        course_pool: A dictionary where the keys are course codes and the values are 
                     course objects. Each course object has a method `get_date()` 
                     that returns the exam date.
        course_code: A string representing the unique code of the course for which 
                     the exam date is to be retrieved.
    
    Returns:
        date: The exam date of the specified course.
    """
    course = course_pool[course_code]
    return course.get_date()

def sameDayDuty(student, date):
    """
    Checks if a specific student has any duties assigned on a given date.
    
    This function iterates through all the duties assigned to the student and 
    checks if any of the duties occur on the specified date. If there is at 
    least one duty on the given date, the function returns True. Otherwise, 
    it returns False.
    
    Args:
        student: An object representing the student, which has a method `get_duties()`.
                 This method returns a list of duty objects.
        date: A date object representing the date to be checked.
    
    Returns:
        bool: True if the student has a duty on the specified date, False otherwise.
    """
    duties = student.get_duties()
    for duty in duties:
        duty_date = duty.get_date()
        if date == duty_date:
            return True
    return False

def removeFromPool(available_pool, available_sorted_keys, student):
    """
    Removes a student's admission number from the available pool and the sorted keys list.
    
    This function takes a dictionary of available students (`available_pool`), a list of 
    sorted keys (`available_sorted_keys`), and a student object (`student`). It removes 
    the student's admission number from both the dictionary and the list if it exists 
    in them. This ensures that the student is no longer considered available for whatever 
    purpose the pool and sorted keys list are being used for.
    
    Args:
        available_pool: A dictionary where keys are student admission numbers and values 
                        are student objects.
        available_sorted_keys: A list of student admission numbers, sorted in a specific 
                               order (e.g., alphabetically or by some other criteria).
        student: An object representing the student, which has a method `get_admission_no()` 
                 that returns the student's admission number.
    
    Returns:
        None
    """
    # Check if the student's admission number is in the available pool
    if student.get_admission_no() in available_pool:
        # Remove the student's admission number from the available pool
        del available_pool[student.get_admission_no()]
    
    # Check if the student's admission number is in the available sorted keys
    if student.get_admission_no() in available_sorted_keys:
        # Remove the student's admission number from the available sorted keys
        available_sorted_keys.remove(student.get_admission_no())


def clearAllotedList(alloted_students_pool, students_available_copy):
    for student in students_available_copy:
        if student in alloted_students_pool:
            del alloted_students_pool[student]

def extract_codes(text):
    """
    Extracts course codes from the given text using a regular expression pattern.
    
    This function defines a regular expression pattern to match course codes,
    which are assumed to be alphanumeric strings starting with letters followed
    by digits and optionally more alphanumeric characters. It finds all matches
    of this pattern in the provided text and returns them as a list.
    
    Args:
        text: A string containing the text from which to extract course codes.
    
    Returns:
        list: A list of strings, each representing a matched course code.
    """
    # Define the regular expression pattern to match codes
    pattern = r'\b[a-zA-Z]+\d+\w*\b'
    
    # Find all matches in the text using the regular expression pattern
    matches = re.findall(pattern, text)
    
    # Return the list of matched codes
    return matches

def custom_sort_key(s):
    """
    Custom sorting key function that extracts and converts the last 5 characters 
    of a string to an integer for sorting purposes.
    
    This function takes a string and returns the last 5 characters of the string 
    as an integer. It can be used as a key for sorting functions to sort strings 
    based on their last 5 characters interpreted as an integer. This is useful 
    for custom sorting requirements where the order depends on a specific substring 
    of the items being sorted.
    
    Args:
        s: A string from which the last 5 characters will be extracted and 
           converted to an integer.
    
    Returns:
        int: The integer value of the last 5 characters of the input string.
    """
    last_five = s[-5:]
    return int(last_five)

def format_name(input_name):
    """
    Formats a given name string by capitalizing the first letter of each word.
    
    This function takes an input string representing a name, splits it into a list 
    of words, and then capitalizes the first letter of each word while converting 
    the rest of the letters to lowercase. The words are then joined back into a 
    single string with spaces in between, resulting in a properly formatted name.
    
    Args:
        input_name: A string representing the name to be formatted.
    
    Returns:
        str: The formatted name with each word capitalized.
    """
    words = input_name.split()  # Split the input string into a list of words
    formatted_name = ' '.join(word.capitalize() for word in words)
    return formatted_name


##########################################################################################################################
################################## Extracting Classroom Data from Classroom Excel File ###################################
##########################################################################################################################

# Dictionary: Room No.(Key) : Strength(Value)
room_capacity = {}

# Relative Path of the Excel File.
file_path = os.path.join(settings.MEDIA_ROOT, 'invigilationFiles', 'Classroom.xlsx')

# Names of the Relevant Columns in the Excel File.
building_col = 'building'
room_col = 'roomNo'
capacity_col = 'capacity'

# Open the Excel file
workbook = openpyxl.load_workbook(filename=file_path)

# Slecting the First(0th - indexed) Worksheet/Sheet in the Workbook/Excel File.
worksheet = workbook.worksheets[0]

# Getting the Column Indices based on the Column Names.
header_row = worksheet[1]
column_names = tuple(cell.value for cell in header_row)

try:
    # Find the indices of the relevant Columns.
    building_col_idx = column_names.index(building_col) + 1
    room_col_idx = column_names.index(room_col) + 1
    capacity_col_idx = column_names.index(capacity_col) + 1
except ValueError as e:
    # Print error if one or more required Columns are not found.
    print("Error: One or more Required Columns not found in the Classroom Excel File!")

# Iterate over rows and extract data from relevant Columns.
for row in worksheet.iter_rows(min_row=2, values_only=True):
    building = row[building_col_idx - 1]
    room = row[room_col_idx - 1]
    capacity = row[capacity_col_idx - 1]

    # Check for missing values in any of the relevant Columns.
    if any(val is None for val in [building, room, capacity]):
            print("Error: Make sure the Columns(Building, Room No., Capacity) are Filled")
            break
    
    # Normalize room name to lowercase and convert capacity to integer.
    room = room.lower().strip()
    capacity = int(capacity)

    # Store the room capacity in a dictionary.
    room_capacity[room] = capacity

#########################################################################################################################
########################## Extracting PhD Students from Student Course Registration Excel File ##########################
#########################################################################################################################

# Dictionary: Admission No.(Key) : Strength(Value)
phd_students = {}

# Relative Path of the Excel file
file_path = os.path.join(settings.MEDIA_ROOT, 'invigilationFiles', 'StudentRegistration.xlsx')

# Names of the relevant Columns in the Excel file
admission_no_col = 'Admission No.'
name_col = 'Name'
email_col = 'Email Id'
course_code_col = 'Course Code'

# Open the Excel file
workbook = openpyxl.load_workbook(filename=file_path)

# Slecting the First(0th - indexed) Worksheet/Sheet in the Workbook/Excel File.
worksheet = workbook.worksheets[0]


# Getting the Column Indices based on the Column Names
header_row = worksheet[1]
column_names = tuple(cell.value for cell in header_row)

try:
    # Find the indices of the relevant Columns.
    admission_no_col_idx = column_names.index(admission_no_col) + 1
    name_col_idx = column_names.index(name_col) + 1
    email_col_idx = column_names.index(email_col) + 1
    course_code_col_idx = column_names.index(course_code_col) + 1
except ValueError as e:
    # Print error if one or more required Columns are not found.
    print("Error: One or more Required Columns not found in the StudentRegistration Excel File!")


# Iterate over rows and extract data from relevant Columns.
for row in worksheet.iter_rows(min_row=2, values_only=True):
    admission_no = row[admission_no_col_idx - 1]
    name = row[name_col_idx - 1]
    email = row[email_col_idx - 1]
    course_code = row[course_code_col_idx - 1]

    # Check for missing values in any of the relevant Columns.
    if any(val is None for val in [admission_no, name, email, course_code]):
            print("Error: Make sure the Columns(Admission No., Name, Email ID, Course Code) are Filled")
            break
    
    # Convert admission number to lowercase
    admission_no = admission_no.lower()

    # Extracting Course Codes
    course_code = course_code.lower()
    # Get list of course codes
    course_code_list = extract_codes(course_code)

    # Check if admission number is not in PhD students
    if admission_no not in phd_students:
        # Add new PhDStudent
        phd_students[admission_no] = PhDStudent(admission_no, name, email)
    
    # Add the enrolled course to the PhDStudent object
    for code in course_code_list:
        phd_students[admission_no].add_enrolled_course(code)

#########################################################################################################################
##################################### Extracting Course List from Exam Date Sheet  ######################################
#########################################################################################################################

# Note: This code assumes that all classrooms allocated for a course belong to the same Building(Instructed by PhD Admin).

# Dictionary: Course Code(Key) : Course Object(Value)
course_pool = {}
# List to store all Course Objects
course_list = []

# Relative Path of the Excel file
file_path = os.path.join(settings.MEDIA_ROOT, 'invigilationFiles', 'ExamDateSheet.xlsx')

# Names of the relevant Columns in the Excel file
date_col = 'Date'
day_col = 'Day'
time_col = 'Time'
accronynm_col = 'Accronynm'
course_code_col = 'Course Code'
strength_col = 'Strength'
room_no_col = 'Room No.'

# Open the Excel file
workbook = openpyxl.load_workbook(filename=file_path)

# Slecting the First(0th - indexed) Worksheet/Sheet in the Workbook/Excel File.
worksheet = workbook.worksheets[0]


# Geting the Column Indices based on the Column Names
header_row = worksheet[1]
column_names = tuple(cell.value for cell in header_row)


try:
    # Find the indices of the relevant Columns.
    date_col_idx = column_names.index(date_col) + 1
    day_col_idx = column_names.index(day_col) + 1
    time_col_idx = column_names.index(time_col) + 1
    accronynm_col_idx = column_names.index(accronynm_col) + 1
    course_code_col_idx = column_names.index(course_code_col) + 1
    strength_col_idx = column_names.index(strength_col) + 1
    room_no_col_idx = column_names.index(room_no_col) + 1
except ValueError as e:
    # Print error if one or more required Columns are not found.
    print("Error: One or more Required Columns not found in the Date Sheet Excel File!")

# Iterate over rows and extract data from relevant Columns.
for row in worksheet.iter_rows(min_row=2, values_only=True):

    date = row[date_col_idx - 1]
    day = row[day_col_idx - 1]
    time = row[time_col_idx - 1]
    accronynm = row[accronynm_col_idx - 1]
    course_code = row[course_code_col_idx - 1]
    strength = row[strength_col_idx - 1]
    room_no = row[room_no_col_idx - 1]

    # Check for missing values in any of the relevant Columns.
    if any(val is None for val in [date, day, time, accronynm, course_code, strength, room_no]):
        print("Error: Make sure the Columns(Date, Day, Time, Accronynm, Course Code, Strength, Room No.) are Filled")
        break

    # Convert Course Code to lowercase.
    course_code = course_code.lower()

    # Making the Course List from the DateSheet.
    course_code_list = extract_codes(course_code)

    # Split room numbers by comma and removing leading and trailing whitespace from each room number.
    room_no = room_no.split(",")
    room_no = [str.strip() for str in room_no]

    # Create a new Course object and add the new course to the course list
    new_course = Course(accronynm, course_code, strength, date, time, day, room_no[:-1], room_capacity, TARatio, room_no[-1])
    course_list.append(new_course)

    # Iterate through course codes list and add the new course to the course pool with each code.
    for code in course_code_list:
        course_pool[code] = new_course


#########################################################################################################################
##################################### Loading Eligible Students for Invilgilation Duty ##################################
#########################################################################################################################

# Dictionary: Admission No.(Key) : PhDStudent Object(Value)
students_pool = {}
students_available = {}

# Relative Path of the Excel file
file_path = os.path.join(settings.MEDIA_ROOT, 'invigilationFiles', 'StudentList.xlsx')

# Names of the relevant Columns in the Excel file
admission_no_col = 'Admission No.'
name_col = 'Name'
email_col = 'Email ID'

# Open the Excel file
workbook = openpyxl.load_workbook(filename=file_path)

# Slecting the First(0th - indexed) Worksheet/Sheet in the Workbook/Excel File.
worksheet = workbook.worksheets[0]

# Geting the Column Indices based on the Column Names
header_row = worksheet[1]
column_names = tuple(cell.value for cell in header_row)

try:
    # Find the indices of the relevant Columns.
    admission_no_col_idx = column_names.index(admission_no_col) + 1
    name_col_idx = column_names.index(name_col) + 1
    email_col_idx = column_names.index(email_col) + 1
except ValueError as e:
    # Print error if one or more required Columns are not found.
    print("Error: One or more Required Columns not found in the Student List Excel File!")

# Iterate over rows and extract data from relevant Columns.
for row in worksheet.iter_rows(min_row=2, values_only=True):
    admission_no = row[admission_no_col_idx - 1]
    name = row[name_col_idx - 1]
    email = row[email_col_idx - 1]
    
    # Check for missing values in any of the relevant Columns.
    if any(val is None for val in [admission_no, name, email]):
        print("Error: Make sure the Columns(Admission No., Name, Email ID) are Filled")
        break

    # Convert the admission number to lowercase.
    admission_no = admission_no.lower()

    # Add a new PhDStudent to phd_students if not already present
    if(admission_no not in phd_students):
        phd_students[admission_no] = PhDStudent(admission_no, name, email)

    # Update students_pool and students_available dictionaries with this student.
    students_pool[admission_no] = phd_students[admission_no]
    students_available[admission_no] = phd_students[admission_no]

#########################################################################################################################
############################################### Extracting Course TAs ###################################################
#########################################################################################################################

# Relative Path of the Excel file
file_path = os.path.join(settings.MEDIA_ROOT, 'invigilationFiles', 'TAList.xlsx')

# Names of the relevant Columns in the Excel file
admission_no_col = 'Admission No.'
email_col = 'Email ID'
course_code_col = 'Course Code'
name_col = 'Name'


# Open the Excel file
workbook = openpyxl.load_workbook(filename=file_path)

# Slecting the First(0th - indexed) Worksheet/Sheet in the Workbook/Excel File.
worksheet = workbook.worksheets[0]

# Geting the Column Indices based on the Column Names
header_row = worksheet[1]
column_names = tuple(cell.value for cell in header_row)


try:
    # Find the indices of the relevant Columns.
    admission_no_col_idx = column_names.index(admission_no_col) + 1
    name_col_idx = column_names.index(name_col) + 1
    email_col_idx = column_names.index(email_col) + 1
    course_code_col_idx = column_names.index(course_code_col) + 1
except ValueError as e:
    # Print error if one or more required Columns are not found.
    print("Error: One or more Required Columns not found in the TAList Excel File!")

# Iterate over rows and extract data from relevant Columns.
for row in worksheet.iter_rows(min_row=2, values_only=True):
    admission_no = row[admission_no_col_idx - 1]
    name = row[name_col_idx - 1]
    email = row[email_col_idx - 1]
    course_code = row[course_code_col_idx - 1]

    # Check for missing values in any of the relevant Columns.
    if any(val is None for val in [admission_no, name, course_code]):
        print("Error: Make sure the Columns(Admission No., Name, Course Code) are Filled - Extracting Course TAs")
        break

    # Convert Admission Number and Course Code to lowercase
    admission_no = admission_no.lower()
    course_code = course_code.lower()

    # Extract course codes from the input
    course_code_list = extract_codes(course_code)

    # Iterate through the list of course codes
    for code in course_code_list:
        # Check if the course code exists in the course pool
        if code in course_pool:
            curr_course = course_pool[code]
            # Check if the admission number exists in the available students pool
            if(admission_no in students_available):
                student = students_available[admission_no]
                # Add the student as a TA for the current course
                curr_course.add_course_ta(student)
                break

#########################################################################################################################
#################################################### Allocation #########################################################
#########################################################################################################################

# Dictionary: Admission No.(Key) : PhDStudent Object(Value)
alloted_students_pool = {}

# Reverse Sorting the Course List based on the Strength of the Course.
course_list = sorted(course_list, key=lambda x: x.get_strength(), reverse=True)

# Sorting the Keys of Available Students in Decending Order. 
available_sorted_keys = sorted(students_available.keys(), key=custom_sort_key, reverse=True)

# Alloting duties to all the TAs for their respective Courses
for course in course_list:
    date = course.get_date()
    
    req_invigilators = course.get_req_invigilator()
    # Check if no Invigilator is required.
    if(req_invigilators == 0):
        continue

    # Get the list of TAs for the course
    course_ta = course.get_course_ta()
    # Iterate through the list of TAs
    for ta in course_ta:
        # Check if the required number of invigilators is met
        if req_invigilators == len(course.get_invigilators()):
            break
        # Check if the TA is in the available students pool
        if ta.get_admission_no() in students_available:
             # Get the list of courses the TA is enrolled in
            ta_course = ta.get_enrolled_courses()
            # Initialize flag to check if TA can be allotted
            can_allot = True
            # Iterate through the TA's enrolled courses and check for exam date conflict
            for ta_course_code in ta_course:
                if (ta_course_code in course_pool) and (getExamDate(course_pool, ta_course_code) == date):
                    can_allot = False
                    break
            # Check if TA can be allotted as invigilator
            if can_allot:
                # Add TA to the course's invigilators
                course.add_invigilators(ta)
                # Remove TA from available pool
                removeFromPool(students_available, available_sorted_keys, ta)
                # add TA to the allotted students pool
                alloted_students_pool[ta.get_admission_no()] = ta
                # Assign the duty to the TA
                ta.add_duty(course)

# Sorting the Keys of Alloted Students Pool in Decending Order. 
alloted_sorted_keys = sorted(alloted_students_pool.keys(), key=custom_sort_key, reverse=True)

for course in course_list:
    date = course.get_date()
    req_invigilators = course.get_req_invigilator()

    # Check if no Invigilator is required.
    if(req_invigilators == 0 or len(course.get_invigilators()) == req_invigilators):
        continue

    full_allotment = False
    while(full_allotment == False):
        for student in available_sorted_keys:
            if req_invigilators == len(course.get_invigilators()):
                full_allotment = True
                break
            student_course = students_available[student].get_enrolled_courses()
            can_allot = True
            for student_course_code in student_course:
                if(((student_course_code in course_pool) and (getExamDate(course_pool, student_course_code) == date)) or (sameDayDuty(students_available[student], date))):
                    can_allot = False
                    break
            if can_allot:
                students_available[student].add_duty(course)
                course.add_invigilators(students_available[student])
                removeFromPool(students_available, available_sorted_keys, students_available[student])
                alloted_students_pool[student] = students_pool[student]

        if(len(available_sorted_keys) == 0):
            students_available = alloted_students_pool.copy()
            available_sorted_keys = sorted(students_available.keys(), key=custom_sort_key, reverse=True)
            clearAllotedList(alloted_students_pool, students_pool)


        alloted_sorted_keys = sorted(alloted_students_pool.keys(), key=custom_sort_key, reverse=True)

        if full_allotment == False:
            for student in alloted_sorted_keys:
                if req_invigilators == len(course.get_invigilators()):
                    full_allotment = True
                    break
                student_course = alloted_students_pool[student].get_enrolled_courses()
                can_allot = True
                for student_course_code in student_course:
                    if(((student_course_code in course_pool) and (getExamDate(course_pool, student_course_code) == date)) or (sameDayDuty(alloted_students_pool[student], date))):
                        can_allot = False
                        break
                if can_allot:
                    alloted_students_pool[student].add_duty(course)
                    course.add_invigilators(alloted_students_pool[student])
                    removeFromPool(alloted_students_pool, alloted_sorted_keys, alloted_students_pool[student])

if(len(available_sorted_keys) == 0):
    clearAllotedList(alloted_students_pool, students_pool)

# #########################################################################################################################
# ############################################### Updating Students List ##################################################
# ######################################################################################################################### 

sorted_courses = sorted(course_list, key=lambda c: (c.get_date(), datetime.strptime(c.get_time().split(' - ')[0], '%I:%M %p')))

# Create an empty dataframe with the specified columns
df = pd.DataFrame(columns=['Date', 'Day', 'Time', 'Course Acronym', 'Course Code', 'Admission No.', 'Name', 'Email ID', 'Room No.', 'Building', 'No. Of. Duties'])
prev_day = ""
prev_course_code = ""
# Iterate through the sorted courses
for course in sorted_courses:
    date = course.get_date().date()
    day = course.get_day()
    time = course.get_time()
    course_acronym = course.get_course_acronym()
    course_code = course.get_course_code()
    room_list = course.get_room_no()
    building = course.get_building()
    strength = course.get_strength()
    distribution = course.get_distribution()


    if course.get_req_invigilator() == 0:
        df = df._append({'Date': date, 'Day': day, 'Time': time, 'Course Acronym': course_acronym, 'Course Code': course_code.upper(),
                        'Admission No.': None, 'Name': None, 'Email ID': None, 'Room No.' : room_list, 'Building' : building, 'No. Of. Duties' : None}, ignore_index=True)
        empty_row = pd.Series([None] * len(df.columns), index=df.columns)
        df = df._append(empty_row, ignore_index=True)
        continue

    distribution_index = 0
    global_count = distribution[distribution_index]
    count = 0
    index = 0

    

    # iterate through the invigilators for the course
    for invigilator in course.get_invigilators():
        
        count += 1
        admission_no = invigilator.get_admission_no().upper()
        name = format_name(invigilator.get_name())
        email = invigilator.get_email().lower()
        num_duties = len(invigilator.get_duties())

        # Check if prev_course_code is different from course_code
        if (prev_course_code != course_code) and (prev_course_code != ""):
            # Append an empty row
            empty_row = pd.Series([None] * len(df.columns), index=df.columns)
            df = df._append(empty_row, ignore_index=True)

        # Check if prev_day is different from day
        if (prev_day != day) and (prev_day != ""):
            # Append two empty rows
            empty_row = pd.Series([None] * len(df.columns), index=df.columns)
            df = df._append(empty_row, ignore_index=True)
                
        

        # add a row to the dataframe for each invigilator
        df = df._append({'Date': date, 'Day': day, 'Time': time, 'Course Acronym': course_acronym, 'Course Code': course_code,
                        'Admission No.': admission_no, 'Name': name, 'Email ID': email, 'Room No.' : room_list[index], 'Building' : building, 'No. Of. Duties' : num_duties}, ignore_index=True)
        
        if count == global_count:
            distribution_index += 1
            if distribution_index < len(distribution):
                global_count = distribution[distribution_index]
                index += 1
                count = 0
        
        prev_day = day
        prev_course_code = course_code

# write the dataframe to an Excel file
df.to_excel('InvigilatorList.xlsx', index=False)