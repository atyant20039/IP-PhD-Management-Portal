import openpyxl
from PhDStudent import PhDStudent
from Course import Course
import pandas as pd
from datetime import datetime
import re
import argparse


# Create an ArgumentParser object
parser = argparse.ArgumentParser()

# Add arguments
parser.add_argument('--TARatio', type=int, help='Value of TARatio')

# Parse the command-line arguments
args = parser.parse_args()

# Access the value of TARatio
TARatio = args.TARatio

print(TARatio)

# To Extract Cource Codes: Returns - List
def extract_codes(text):
    # Define the regular expression pattern to match codes
    pattern = r'\b[a-zA-Z]+\d+\w*\b'

    # Find all matches in the text using the regular expression pattern
    matches = re.findall(pattern, text)

    return matches

def hasExam(course_pool, course_code):
    course = course_pool[course_code]
    return course.get_date()

def removeFromPool(available_pool, available_sorted_keys, student):
    if student.get_admission_no() in available_pool:
        del available_pool[student.get_admission_no()]
    if student.get_admission_no() in available_sorted_keys:
        available_sorted_keys.remove(student.get_admission_no())

def sameDayDuty(student, date):
    duties = student.get_duties()
    for duty in duties:
        duty_date = duty.get_date()
        if date == duty_date:
            return True
    return False

def clearAllotedList(alloted_students_pool, students_available_copy):
    for student in students_available_copy:
        if student in alloted_students_pool:
            del alloted_students_pool[student]

# def get_num_invigilator(strength):
#         result = strength / 30
#         if strength >= 150 or strength % 30 >= 10:
#             return math.ceil(result)
#         else:
#             return math.floor(result)
        
# Define a custom sorting function based on the last 5 characters
def custom_sort_key(s):
    last_five = s[-5:]  # Get the last 5 characters
    return int(last_five)  # Convert to integer and negate for descending order

def format_name(input_name):
    words = input_name.split()  # Split the input string into a list of words
    formatted_name = ' '.join(word.capitalize() for word in words)
    return formatted_name



##########################################################################################################################
################################## Extracting Classroom Data from Classroom Excel File ###################################
##########################################################################################################################

room_capacity = {}

# relative Path of the Excel File.
file_path = 'Classroom.xlsx'

# Names of the relevant Columns in the Excel File.
building_col = 'building'
room_col = 'roomNo'
capacity_col = 'capacity'

# Open the Excel file
workbook = openpyxl.load_workbook(filename=file_path)

# Slecting the First(0th - indexed) Worksheet/Sheet in the Workbook/Excel File.
worksheet = workbook.worksheets[0]

# Geting the Column Indices based on the Column Names
header_row = worksheet[1]
column_names = tuple(cell.value for cell in header_row)

try:
    building_col_idx = column_names.index(building_col) + 1
    room_col_idx = column_names.index(room_col) + 1
    capacity_col_idx = column_names.index(capacity_col) + 1
except ValueError as e:
    print("Error: One or more Required Columns not found in the Excel File!")


# Iterate over rows and extract data from relevant Columns.
for row in worksheet.iter_rows(min_row=2, values_only=True):
    building = row[building_col_idx - 1]
    room = row[room_col_idx - 1]
    capacity = row[capacity_col_idx - 1]

    if any(val is None for val in [building, room, capacity]):
            print("Error: Make sure the Columns(Building, Room No., Capacity) are Filled")
            break
    
    room = room.lower()
    capacity = int(capacity)

    room_capacity[room] = capacity

#########################################################################################################################
########################## Extracting PhD Students from Student Course Registration Excel File ##########################
#########################################################################################################################

# List of Students from "StudentRegistration" Excel File.
phd_students = {}

# Relative Path of the Excel file
file_path = 'StudentRegistration.xlsx'

# Names of the relevant Columns in the Excel file
admission_no_col = 'Admission No.'
name_col = 'Name'
email_col = 'Email Id'
course_code_col = 'Course Code'

# Open the Excel file
workbook = openpyxl.load_workbook(filename=file_path)

# Slecting the First(0th - indexed) Worksheet/Sheet in the Workbook/Excel File.
worksheet = workbook.worksheets[0]


# Geting the Column Indices based on the Column Names
header_row = worksheet[1]
column_names = tuple(cell.value for cell in header_row)

try:
    admission_no_col_idx = column_names.index(admission_no_col) + 1
    name_col_idx = column_names.index(name_col) + 1
    email_col_idx = column_names.index(email_col) + 1
    course_code_col_idx = column_names.index(course_code_col) + 1
except ValueError as e:
    print("Error: One or more Required Columns not found in the Excel File!")


# Iterate over rows and extract data from relevant Columns.
for row in worksheet.iter_rows(min_row=2, values_only=True):
    admission_no = row[admission_no_col_idx - 1]
    name = row[name_col_idx - 1]
    email = row[email_col_idx - 1]
    course_code = row[course_code_col_idx - 1]


    if any(val is None for val in [admission_no, name, email, course_code]):
            print("Error: Make sure the Columns(Admission No., Name, Email ID, Course Code) are Filled")
            break
    
    admission_no = admission_no.lower()

    # Extracting Course Codes
    course_code = course_code.lower()
    course_code_list = extract_codes(course_code)

    if admission_no not in phd_students:
            phd_students[admission_no] = PhDStudent(admission_no, name, email)
    
    # Add the enrolled course to the PhDStudent object
    # phd_students[admission_no].add_enrolled_course(course_code)
    for code in course_code_list:
        phd_students[admission_no].add_enrolled_course(code)

#########################################################################################################################
##################################### Extracting Course List from Exam Date Sheet  ######################################
#########################################################################################################################
        
course_pool = {}
course_list = []

# Relative Path of the Excel file
file_path = 'ExamDateSheet.xlsx'

# Names of the relevant Columns in the Excel file
date_col = 'Date'
day_col = 'Day'
time_col = 'Time'
accronynm_col = 'Accronynm'
course_code_col = 'Course Code'
strength_col = 'Strength'
room_no_col = 'Room No.'
# building_col = 'Building'

# Open the Excel file
workbook = openpyxl.load_workbook(filename=file_path)

# Slecting the First(0th - indexed) Worksheet/Sheet in the Workbook/Excel File.
worksheet = workbook.worksheets[0]


# Geting the Column Indices based on the Column Names
header_row = worksheet[1]
column_names = tuple(cell.value for cell in header_row)


try:
    date_col_idx = column_names.index(date_col) + 1
    day_col_idx = column_names.index(day_col) + 1
    time_col_idx = column_names.index(time_col) + 1
    accronynm_col_idx = column_names.index(accronynm_col) + 1
    course_code_col_idx = column_names.index(course_code_col) + 1
    strength_col_idx = column_names.index(strength_col) + 1
    room_no_col_idx = column_names.index(room_no_col) + 1
    # buiding_col_idx = column_names.index(building_col) + 1

except ValueError as e:
    print("Error: One or more Required Columns not found in the Date Sheet Excel File!")

count = 1

# Iterate over rows and extract data from relevant Columns.
for row in worksheet.iter_rows(min_row=2, values_only=True):

    date = row[date_col_idx - 1]
    day = row[day_col_idx - 1]
    time = row[time_col_idx - 1]
    accronynm = row[accronynm_col_idx - 1]
    course_code = row[course_code_col_idx - 1]
    strength = row[strength_col_idx - 1]
    room_no = row[room_no_col_idx - 1]
    # building = row[building_col_idx - 1]

    if any(val is None for val in [date, day, time, accronynm, course_code, strength, room_no]):
        print("Error: Make sure the Columns(Date, Day, Time, Accronynm, Course Code, Strength, Room No.) are Filled")
        break

    course_code = course_code.lower()
    # Making the Course List from the DateSheet.
    course_code_list = extract_codes(course_code)
    course_code_original = course_code

    room_no = room_no.split(",")
    room_no = [str.strip() for str in room_no]

    # room_no = extract_codes(room_no)
    # print(room_no)

    new_course = Course(accronynm, course_code_original, strength, date, time, day, room_no, room_capacity, TARatio) # Change Course Structure
    
    count = count + 1
    course_list.append(new_course)
    for code in course_code_list:
        course_pool[code] = new_course


#########################################################################################################################
##################################### Loading Eligible Students for Invilgilation Duty ##################################
#########################################################################################################################

students_pool = {}
students_available = {}


# Relative Path of the Excel file
file_path = 'StudentList.xlsx'

# Names of the relevant Columns in the Excel file
admission_no_col = 'Admission No.'
name_col = 'Name'
email_col = 'Email ID'

# Open the Excel file
workbook = openpyxl.load_workbook(filename=file_path)

# Slecting the First(0th - indexed) Worksheet/Sheet in the Workbook/Excel File.
worksheet = workbook.worksheets[0]

# Geting the Column Indices based on the Column Names
header_row = worksheet[1]
column_names = tuple(cell.value for cell in header_row)


try:
    admission_no_col_idx = column_names.index(admission_no_col) + 1
    name_col_idx = column_names.index(name_col) + 1
    email_col_idx = column_names.index(email_col) + 1
except ValueError as e:
    print("Error: One or more Required Columns not found in the Excel File!")

# Iterate over rows and extract data from relevant Columns.
for row in worksheet.iter_rows(min_row=2, values_only=True):
    admission_no = row[admission_no_col_idx - 1]
    name = row[name_col_idx - 1]
    email = row[email_col_idx - 1]
    
    
    if any(val is None for val in [admission_no, name, email]):
        print("Error: Make sure the Columns(Admission No., Name, Email ID) are Filled")
        break

    admission_no = admission_no.lower()

    if(admission_no not in phd_students):
        phd_students[admission_no] = PhDStudent(admission_no, name, email)

    students_pool[admission_no] = phd_students[admission_no]
    students_available[admission_no] = phd_students[admission_no]
        

students_available_copy = students_available.copy()

#########################################################################################################################
############################################### Extracting Course TAs ###################################################
#########################################################################################################################

# Relative Path of the Excel file
file_path = 'TAList.xlsx'

# Names of the relevant Columns in the Excel file
admission_no_col = 'Admission No.'
email_col = 'Email ID'
course_code_col = 'Course Code'
name_col = 'Name'


# Open the Excel file
workbook = openpyxl.load_workbook(filename=file_path)

# Slecting the First(0th - indexed) Worksheet/Sheet in the Workbook/Excel File.
worksheet = workbook.worksheets[0]

# Geting the Column Indices based on the Column Names
header_row = worksheet[1]
column_names = tuple(cell.value for cell in header_row)


try:
    admission_no_col_idx = column_names.index(admission_no_col) + 1
    name_col_idx = column_names.index(name_col) + 1
    email_col_idx = column_names.index(email_col) + 1
    course_code_col_idx = column_names.index(course_code_col) + 1
except ValueError as e:
    print("Error: One or more Required Columns not found in the Excel File!")

# Iterate over rows and extract data from relevant Columns.
for row in worksheet.iter_rows(min_row=2, values_only=True):
    admission_no = row[admission_no_col_idx - 1]
    name = row[name_col_idx - 1]
    email = row[email_col_idx - 1]
    course_code = row[course_code_col_idx - 1]

    
    if any(val is None for val in [admission_no, name, course_code]):
        print("Error: Make sure the Columns(Admission No., Name, Course Code) are Filled - Extracting Course TAs")
        break

    

    admission_no = admission_no.lower()
    course_code = course_code.lower()
    course_code_list = extract_codes(course_code)
    for code in course_code_list:
        if code in course_pool:
            curr_course = course_pool[code]
            if(admission_no in students_pool):
                student = students_pool[admission_no]
                curr_course.add_course_ta(student)
                break


#########################################################################################################################
#################################################### Allocation #########################################################
#########################################################################################################################

alloted_students_pool = {}

# Reverse Sorting the Course List based on the Strength of the Course.
course_list = sorted(course_list, key=lambda x: x.get_strength(), reverse=True)

# Sorting the Keys of Available Students in Decending Order. 
available_sorted_keys = sorted(students_available.keys(), key=custom_sort_key, reverse=True)

# Alloting duties to all the TAs for their respective Courses
for course in course_list:
    date = course.get_date()
    
    req_invigilators = course.get_req_invigilator()
    # Check if no Invigilator is required.
    if(req_invigilators == 0):
        continue

    course_ta = course.get_course_ta()
    for ta in course_ta:
        if req_invigilators == len(course.get_invigilators()):
            break
        if ta.get_admission_no() in students_pool:
            ta_course = ta.get_enrolled_courses()
            can_allot = True
            for ta_course_code in ta_course:
                if (ta_course_code in course_pool) and (hasExam(course_pool, ta_course_code) == date):
                    can_allot = False
                    break
            if can_allot:
                course.add_invigilators(ta)
                removeFromPool(students_available, available_sorted_keys, ta)
                alloted_students_pool[ta.get_admission_no()] = ta
                ta.add_duty(course)

# Sorting the Keys of Alloted Students Pool in Decending Order. 
alloted_sorted_keys = sorted(alloted_students_pool.keys(), key=custom_sort_key, reverse=True)

for course in course_list:
    date = course.get_date()
    req_invigilators = course.get_req_invigilator()

    # Check if no Invigilator is required.
    if(req_invigilators == 0 or len(course.get_invigilators()) == req_invigilators):
        continue

    full_allotment = False
    while(full_allotment == False):
        for student in available_sorted_keys:
            if req_invigilators == len(course.get_invigilators()):
                full_allotment = True
                break
            student_course = students_available[student].get_enrolled_courses()
            can_allot = True
            for student_course_code in student_course:
                if(((student_course_code in course_pool) and (hasExam(course_pool, student_course_code) == date)) or (sameDayDuty(students_available[student], date))):
                    can_allot = False
                    break
            if can_allot:
                students_available[student].add_duty(course)
                course.add_invigilators(students_available[student])
                removeFromPool(students_available, available_sorted_keys, students_available[student])
                alloted_students_pool[student] = students_pool[student]

        if(len(available_sorted_keys) == 0):
            students_available = alloted_students_pool.copy()
            available_sorted_keys = sorted(students_available.keys(), key=custom_sort_key, reverse=True)
            clearAllotedList(alloted_students_pool, students_available_copy)


        alloted_sorted_keys = sorted(alloted_students_pool.keys(), key=custom_sort_key, reverse=True)

        if full_allotment == False:
            for student in alloted_sorted_keys:
                if req_invigilators == len(course.get_invigilators()):
                    full_allotment = True
                    break
                student_course = alloted_students_pool[student].get_enrolled_courses()
                can_allot = True
                for student_course_code in student_course:
                    if(((student_course_code in course_pool) and (hasExam(course_pool, student_course_code) == date)) or (sameDayDuty(alloted_students_pool[student], date))):
                        can_allot = False
                        break
                if can_allot:
                    alloted_students_pool[student].add_duty(course)
                    course.add_invigilators(alloted_students_pool[student])
                    removeFromPool(alloted_students_pool, alloted_sorted_keys, alloted_students_pool[student])

if(len(available_sorted_keys) == 0):
    clearAllotedList(alloted_students_pool, students_available_copy)

# #########################################################################################################################
# ############################################### Updating Students List ##################################################
# ######################################################################################################################### 

sorted_courses = sorted(course_list, key=lambda c: (c.get_date(), datetime.strptime(c.get_time().split(' - ')[0], '%I:%M %p')))

# create an empty dataframe with the specified columns
df = pd.DataFrame(columns=['Date', 'Day', 'Time', 'Course Acronym', 'Course Code', 'Admission No.', 'Name', 'Email ID', 'Room No.', 'Building', 'No. Of. Duties'])
prev_day = ""
prev_course_code = ""
# iterate through the sorted courses
for course in sorted_courses:
    date = course.get_date().date()
    day = course.get_day()
    time = course.get_time()
    course_acronym = course.get_course_acronym()
    course_code = course.get_course_code()
    room_list = course.get_room_no()
    building = course.get_building()
    strength = course.get_strength()
    distribution = course.get_distribution()

    if course.get_req_invigilator() == 0:
        continue

    distribution_index = 0
    global_count = distribution[distribution_index]
    count = 0
    index = 0

    # iterate through the invigilators for the course
    for invigilator in course.get_invigilators():
        count += 1
        admission_no = invigilator.get_admission_no().upper()
        name = format_name(invigilator.get_name())
        email = invigilator.get_email().lower()
        num_duties = len(invigilator.get_duties())

        # Check if prev_course_code is different from course_code
        if (prev_course_code != course_code) and (prev_course_code != ""):
            # Append an empty row
            empty_row = pd.Series([None] * len(df.columns), index=df.columns)
            df = df._append(empty_row, ignore_index=True)

        # Check if prev_day is different from day
        if (prev_day != day) and (prev_day != ""):
            # Append two empty rows
            empty_row = pd.Series([None] * len(df.columns), index=df.columns)
            df = df._append(empty_row, ignore_index=True)
                
        

        # add a row to the dataframe for each invigilator
        df = df._append({'Date': date, 'Day': day, 'Time': time, 'Course Acronym': course_acronym, 'Course Code': course_code.upper(),
                        'Admission No.': admission_no, 'Name': name, 'Email ID': email, 'Room No.' : room_list[index], 'Building' : building, 'No. Of. Duties' : num_duties}, ignore_index=True)
        
        if count == global_count:
            distribution_index += 1
            if distribution_index < len(distribution):
                global_count = distribution[distribution_index]
                index += 1
                count = 0

        prev_day = day
        prev_course_code = course_code

# write the dataframe to an Excel file
df.to_excel('InvigilatorList.xlsx', index=False)
