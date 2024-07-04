import os
import re
import subprocess

from django.conf import settings
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from django.db.models import Q
from django.http import HttpResponse
from django_filters.rest_framework import DjangoFilterBackend
from openpyxl import Workbook, load_workbook
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.worksheet.datavalidation import DataValidation
from rest_framework import status
from rest_framework.filters import OrderingFilter, SearchFilter
from rest_framework.mixins import CreateModelMixin, ListModelMixin
from rest_framework.parsers import MultiPartParser
from rest_framework.response import Response
from rest_framework.viewsets import (GenericViewSet, ModelViewSet,
                                     ReadOnlyModelViewSet, ViewSet)

from .models import *
from .pagination import NoPagination
from .serializers import *


class StudentRegistrationTemplateViewSet(ListModelMixin, GenericViewSet):
    parser_classes = [MultiPartParser]

    def get_queryset(self):
        return []
    
    # To get StudentRegistration Template
    def list(self, request, *args, **kwargs):
        # Define the file path
        file_path = os.path.join(settings.BASE_DIR, 'templates', 'StudentRegistrationTemplate.xlsx')

        # Return the Excel file as a response for download
        with open(file_path, 'rb') as file:
            response = HttpResponse(
                file.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="StudentRegistrationTemplate.xlsx"'
            return response
        
    def create(self, request, *args, **kwargs):
        if 'file' not in request.data:
            return Response({'error': 'No File Uploaded'}, status=status.HTTP_400_BAD_REQUEST)
        
        uploaded_file = request.data['file']

        if not uploaded_file.name.endswith('.xlsx'):
            return Response({'error': 'Invalid File Format. Please upload an Excel File (.xlsx)'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            # Save the file first
            file_content = uploaded_file.read()
            temp_file_path = os.path.join(settings.MEDIA_ROOT, 'temp', uploaded_file.name)
            os.makedirs(os.path.dirname(temp_file_path), exist_ok=True)

            with open(temp_file_path, 'wb') as temp_file:
                temp_file.write(file_content)


            workbook = load_workbook(uploaded_file)
            sheet = workbook.active

            sheet_headers = [cell.value for cell in sheet[1]]
            required_columns = [
                'Admission No.',
                'Name',
                'Email ID',
                'Course Code'
            ]

            if not all(col in sheet_headers for col in required_columns):
                return Response({'error': 'Missing required Columns in the Excel File'}, status=status.HTTP_400_BAD_REQUEST)
            
            invalid_rows = []
            admission_no_index = sheet_headers.index('Admission No.')
            # email_id_index = sheet_headers.index('Email ID')
            course_code_index = sheet_headers.index('Course Code')

            admission_no_pattern = re.compile(r'^(phd\d{5}|mt\d{5})$')
            # email_pattern = re.compile(r'^[a-zA-Z0-9._%+-]+@iiitd\.ac\.in$')
            course_code_pattern = re.compile(r'^[a-zA-Z0-9\s/]+$')

            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                reasons = []
                if all(cell is None for cell in row):
                    continue
                elif any(cell is None for cell in row[:4]  or any(str(cell).isspace() for cell in row[:4])):
                    reasons.append('Makesure all the Cells are filled!')
                    invalid_rows.append({
                        'row': row_idx,
                        'reasons': reasons
                    })
                    continue

                course_code = row[course_code_index]
                admission_no = row[admission_no_index]
                # email_id = row[email_id_index]

                if isinstance(row[course_code_index], str):
                    course_code = course_code.strip() 
                if isinstance(row[admission_no_index], str):
                    admission_no = admission_no.strip()
                # if isinstance(row[email_id_index], str):
                #     email_id = email_id.strip()

                if not isinstance(row[admission_no_index], str) or not admission_no_pattern.match(admission_no.lower()):
                    reasons.append(f'[{admission_no}] Invalid Admission No. Format.')

                # if not isinstance(row[email_id_index], str) or not email_pattern.match(email_id.lower().strip()):
                #     reasons.append(f'[{email_id}] Invalid Email Id Format. Please use IIITD Emails only')

                if not isinstance(row[course_code_index], str) or not course_code_pattern.match(course_code):
                    reasons.append(f'[{course_code}]: Invalid Format. Make sure multiple Course Codes are "/" Separated. Use (A-Z), (a-z), (0-9) characters only.')
                
                if reasons:
                    invalid_rows.append({
                        'row': row_idx,
                        'reasons': reasons
                    })

            if invalid_rows:
                return Response({'error': 'Invalid Data in Excel File', 'invalid_rows': invalid_rows}, status=status.HTTP_400_BAD_REQUEST)
            else:
                # Save the validated file to the MEDIA_ROOT directory
                final_file_path = os.path.join(settings.MEDIA_ROOT, 'invigilationFiles', 'StudentRegistration.xlsx')
                os.makedirs(os.path.dirname(final_file_path), exist_ok=True)

                # Remove the file if it already exists
                if default_storage.exists(final_file_path):
                    default_storage.delete(final_file_path)

                # Save the validated content
                default_storage.save(final_file_path, ContentFile(file_content))
                return Response({'message': 'File Uploaded and Validated successfully.'}, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    


class StudentListTemplateViewSet(ListModelMixin, GenericViewSet):
    parser_classes = [MultiPartParser]

    def get_queryset(self):
        return []
    
    # To get StudentList Template
    def list(self, request, *args, **kwargs):
        # Define the file path
        file_path = os.path.join(settings.BASE_DIR, 'templates', 'StudentListTemplate.xlsx')

        # Return the Excel file as a response for download
        with open(file_path, 'rb') as file:
            response = HttpResponse(
                file.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="StudentListTemplate.xlsx"'
            return response
        
    def create(self, request, *args, **kwargs):
        if 'file' not in request.data:
            return Response({'error': 'No File Uploaded'}, status=status.HTTP_400_BAD_REQUEST)
        
        uploaded_file = request.data['file']

        if not uploaded_file.name.endswith('.xlsx'):
            return Response({'error': 'Invalid File Format. Please upload an Excel File (.xlsx)'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            # Save the file first
            file_content = uploaded_file.read()
            temp_file_path = os.path.join(settings.MEDIA_ROOT, 'temp', uploaded_file.name)
            os.makedirs(os.path.dirname(temp_file_path), exist_ok=True)

            with open(temp_file_path, 'wb') as temp_file:
                temp_file.write(file_content)

            workbook = load_workbook(uploaded_file)
            sheet = workbook.active

            sheet_headers = [cell.value for cell in sheet[1]]
            required_columns = [
                'Admission No.',
                'Name',
                'Email ID'
            ]

            if not all(col in sheet_headers for col in required_columns):
                return Response({'error': 'Missing required Columns in the Excel File'}, status=status.HTTP_400_BAD_REQUEST)
            
            invalid_rows = []
            admission_no_index = sheet_headers.index('Admission No.')
            # email_id_index = sheet_headers.index('Email ID')

            admission_no_pattern = re.compile(r'^(phd\d{5}|mt\d{5})$')
            # email_pattern = re.compile(r'^[a-zA-Z0-9._%+-]+@iiitd\.ac\.in$')

            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                reasons = []
                if all(cell is None for cell in row):
                    continue
                elif any(cell is None for cell in row[:3]  or any(str(cell).isspace() for cell in row[:3])):
                    reasons.append('Makesure all the Cells are Filled!')
                    invalid_rows.append({
                        'row': row_idx,
                        'reasons': reasons
                    })
                    continue

                admission_no = row[admission_no_index]

                if isinstance(row[admission_no_index], str):
                    admission_no = admission_no.strip()

                # if isinstance(row[email_id_index], str):
                #     email_id = row[email_id_index].strip()

                if not isinstance(row[admission_no_index], str) or not admission_no_pattern.match(admission_no.lower().strip()):
                    reasons.append(f'[{admission_no}] Invalid Admission No. Format.')

                # if not isinstance(row[email_id_index], str) or not email_pattern.match(email_id.lower().strip()):
                #     reasons.append(f'[{email_id}] Invalid Email Id Format.')
                
                if reasons:
                    invalid_rows.append({
                        'row': row_idx,
                        'reasons': reasons
                    })

            if invalid_rows:
                return Response({'error': 'Invalid Data in Excel File', 'invalid_rows': invalid_rows}, status=status.HTTP_400_BAD_REQUEST)
            else:
                # Save the validated file to the MEDIA_ROOT directory
                final_file_path = os.path.join(settings.MEDIA_ROOT, 'invigilationFiles', 'StudentList.xlsx')
                os.makedirs(os.path.dirname(final_file_path), exist_ok=True)

                # Remove the file if it already exists
                if default_storage.exists(final_file_path):
                    default_storage.delete(final_file_path)

                # Save the validated content
                default_storage.save(final_file_path, ContentFile(file_content))
                return Response({'message': 'File Uploaded and Validated Successfully.'}, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        

class TATemplateViewSet(ListModelMixin, GenericViewSet):
    parser_classes = [MultiPartParser]

    def get_queryset(self):
        return []
    
    # To get TA Template
    def list(self, request, *args, **kwargs):
        # Define the file path
        file_path = os.path.join(settings.BASE_DIR, 'templates', 'TATemplate.xlsx')

        # Return the Excel file as a response for download
        with open(file_path, 'rb') as file:
            response = HttpResponse(
                file.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="TATemplate.xlsx"'
            return response
        
    def create(self, request, *args, **kwargs):
        if 'file' not in request.data:
            return Response({'error': 'No File Uploaded'}, status=status.HTTP_400_BAD_REQUEST)
        
        uploaded_file = request.data['file']

        if not uploaded_file.name.endswith('.xlsx'):
            return Response({'error': 'Invalid File Format. Please upload an Excel File (.xlsx)'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            # Save the file first
            file_content = uploaded_file.read()
            temp_file_path = os.path.join(settings.MEDIA_ROOT, 'temp', uploaded_file.name)
            os.makedirs(os.path.dirname(temp_file_path), exist_ok=True)

            with open(temp_file_path, 'wb') as temp_file:
                temp_file.write(file_content)

            workbook = load_workbook(uploaded_file)
            sheet = workbook.active

            sheet_headers = [cell.value for cell in sheet[1]]
            required_columns = [
                'Admission No.',
                'Name',
                'Course Code'
            ]

            if not all(col in sheet_headers for col in required_columns):
                return Response({'error': 'Missing required Columns in the Excel File'}, status=status.HTTP_400_BAD_REQUEST)
            
            invalid_rows = []
            course_code_index = sheet_headers.index('Course Code')
            admission_no_index = sheet_headers.index('Admission No.')

            admission_no_pattern = re.compile(r'^(phd\d{5}|mt\d{5})$')
            course_code_pattern = re.compile(r'^[a-zA-Z0-9\s/]+$')
            course_code_pattern_sec = r'[A-Z]{3}\d{3}'

            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                reasons = []
                if all(cell is None for cell in row):
                    continue
                elif any(cell is None for cell in row[:3] or any(str(cell).isspace() for cell in row[:3])):
                    reasons.append('Makesure all the Cells are Filled!')
                    invalid_rows.append({
                        'row': row_idx,
                        'reasons': reasons
                    })
                    continue

                
                course_code = row[course_code_index]
                admission_no = row[admission_no_index]

                if isinstance(row[course_code_index], str):
                    course_code = course_code.strip()
                    course_code = re.findall(course_code_pattern_sec, course_code)
                    course_code = '/'.join(course_code) 
                if isinstance(row[admission_no_index], str):
                    admission_no = admission_no.strip()

                if not isinstance(course_code, str) or not course_code_pattern.match(course_code):
                    reasons.append(f'[{course_code}] Invalid Format. Make sure multiple Course Codes are "/" Separated. Use (A-Z), (a-z), (0-9) characters only.')

                if not isinstance(admission_no, str) or not admission_no_pattern.match(admission_no.lower()):
                    reasons.append(f'[{admission_no}] Invalid Admission No. Format.')
                
                if reasons:
                    invalid_rows.append({
                        'row': row_idx,
                        'reasons': reasons
                    })

            if invalid_rows:
                return Response({'error': 'Invalid Data in Excel File', 'invalid_rows': invalid_rows}, status=status.HTTP_400_BAD_REQUEST)
            else:
                # Save the validated file to the MEDIA_ROOT directory
                final_file_path = os.path.join(settings.MEDIA_ROOT, 'invigilationFiles', 'TAList.xlsx')
                os.makedirs(os.path.dirname(final_file_path), exist_ok=True)

                # Remove the file if it already exists
                if default_storage.exists(final_file_path):
                    default_storage.delete(final_file_path)

                # Save the validated content
                default_storage.save(final_file_path, ContentFile(file_content))
                return Response({'message': 'File Uploaded and Validated Successfully.'}, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        

class ExamDateSheetTemplateViewSet(ListModelMixin, GenericViewSet):
    parser_classes = [MultiPartParser]

    def get_queryset(self):
        return []
    
    # To get DateSheetTemplate
    def list(self, request, *args, **kwargs):
        # Define the file path
        file_path = os.path.join(settings.BASE_DIR, 'templates', 'DateSheetTemplate.xlsx')

        # Return the Excel file as a response for download
        with open(file_path, 'rb') as file:
            response = HttpResponse(
                file.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="DateSheetTemplate.xlsx"'
            return response
        
    def create(self, request, *args, **kwargs):
        if 'file' not in request.data:
            return Response({'error': 'No File Uploaded'}, status=status.HTTP_400_BAD_REQUEST)
        
        uploaded_file = request.data['file']

        if not uploaded_file.name.endswith('.xlsx'):
            return Response({'error': 'Invalid File Format. Please upload an Excel File (.xlsx)'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            # Save the file first
            file_content = uploaded_file.read()
            temp_file_path = os.path.join(settings.MEDIA_ROOT, 'temp', uploaded_file.name)
            os.makedirs(os.path.dirname(temp_file_path), exist_ok=True)

            with open(temp_file_path, 'wb') as temp_file:
                temp_file.write(file_content)

            workbook = load_workbook(uploaded_file)
            sheet = workbook.active

            sheet_headers = [cell.value for cell in sheet[1]]
            required_columns = [
                'Date',
                'Day',
                'Time',
                'Accronynm',
                'Course Code',
                'Strength',
                'Room No.',
                'Building'
            ]

            if not all(col in sheet_headers for col in required_columns):
                return Response({'error': 'Missing required Columns in the Excel File'}, status=status.HTTP_400_BAD_REQUEST)

            invalid_rows = []
            course_code_index = sheet_headers.index('Course Code')
            room_no_index = sheet_headers.index('Room No.')
            strength_index = sheet_headers.index('Strength')

            course_code_pattern = re.compile(r'^[a-zA-Z0-9\s/]+$')
            room_no_pattern = re.compile(r'^[a-zA-Z0-9\s]+(?:\s*,\s*[a-zA-Z0-9\s]+)*$')
            codes = []

            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                reasons = []
                if all(cell is None for cell in row):
                    continue
                elif any(cell is None for cell in row[:7]) or any(str(cell).isspace() for cell in row[:7]):
                    reasons.append('Makesure all the Cells are Filled!')
                    invalid_rows.append({
                        'row': row_idx,
                        'reasons': reasons
                    })
                    continue

                course_code = row[course_code_index]
                room_no = row[room_no_index]
                strength = row[strength_index]

                if isinstance(row[course_code_index], str):
                    course_code = row[course_code_index].strip()
                
                if isinstance(row[room_no_index], str):
                    room_no = row[room_no_index].strip()

                if not type(strength) == int:
                    reasons.append(f'[{strength}] Invalid Value: Strength should be numeric.')

                if not isinstance(row[course_code_index], str) or not course_code_pattern.match(course_code):
                    reasons.append(f'[{course_code}] Invalid Format. Make sure multiple Course Codes are "/" Separated. Use (A-Z), (a-z), (0-9) characters only.')
                else:
                    for code in course_code.split('/'):
                        codes.append(code.strip())
                
                if not isinstance(row[room_no_index], str) or not room_no_pattern.match(room_no):
                    reasons.append(f'[{room_no}] Invalid Format. Make sure multiple Room Nos. are "," Separated. Use (A-Z), (a-z), (0-9) characters only.')
                else:
                    room_no_list = room_no.split(",")
                    room_no_list = [str.strip() for str in room_no_list]

                    # Check if every classroom in the list exists in the database
                    for room in room_no_list:
                        if len(room) > 0 and not (Classroom.objects.filter(roomNo=room).exists()):
                            reasons.append(f'Classroom {room} does not exist. Kindly check the Classroom Table.')                    

                if reasons:
                    invalid_rows.append({
                        'row': row_idx,
                        'reasons': reasons
                    })

            if invalid_rows:
                return Response({'error': 'Invalid Data in Excel File', 'invalid_rows': invalid_rows}, status=status.HTTP_400_BAD_REQUEST)
            else:
                # Save the validated file to the MEDIA_ROOT directory
                final_file_path = os.path.join(settings.MEDIA_ROOT, 'invigilationFiles', 'ExamDateSheet.xlsx')
                os.makedirs(os.path.dirname(final_file_path), exist_ok=True)

                # Remove the file if it already exists
                if default_storage.exists(final_file_path):
                    default_storage.delete(final_file_path)

                # Save the validated content
                default_storage.save(final_file_path, ContentFile(file_content))
                return Response({'message': 'File Uploaded and Validated Successfully.'}, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class StudentViewSet(ModelViewSet):
    queryset = Student.objects.all()
    serializer_class = StudentSerializer
    lookup_field = 'rollNumber'
    lookup_url_kwarg = 'rollNumber'
    filter_backends = [DjangoFilterBackend, SearchFilter]
    filterset_fields = ['name', 'emailId', 'rollNumber', 'studentStatus', 'fundingType', 'gender', 'department', 'region', 'batch', 'admissionThrough']
    search_fields = ['$name', '$emailId', '$rollNumber']

class AllStudentViewSet(ModelViewSet):
    queryset = Student.objects.all()
    serializer_class = StudentSerializer
    lookup_field = 'rollNumber'
    lookup_url_kwarg = 'rollNumber'
    pagination_class = NoPagination
    filter_backends = [DjangoFilterBackend, SearchFilter]
    filterset_fields = ['name', 'emailId', 'rollNumber', 'studentStatus', 'fundingType', 'gender', 'department', 'region', 'batch', 'admissionThrough']
    search_fields = ['$name', '$emailId', '$rollNumber']

class StudentTableViewSet(ModelViewSet):
    queryset = Student.objects.all()
    serializer_class = StudentTableSerializer
    lookup_field = 'rollNumber'
    lookup_url_kwarg = 'rollNumber'
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['name', 'emailId', 'rollNumber', 'studentStatus', 'gender', 'department', 'batch', 'admissionThrough', 'region', 'fundingType', 'yearOfLeaving']
    search_fields = ['$name', '$emailId', '$rollNumber', '$advisor_set__advisor1__name', '$advisor_set__advisor2__name', '$advisor_set__coadvisor__name']

    def get_queryset(self):
        queryset = super().get_queryset()
        advisor_name = self.request.query_params.get('advisor', None)
        if advisor_name:
            queryset = queryset.filter(
                Q(advisor_set__advisor1__name__icontains=advisor_name) |
                Q(advisor_set__advisor2__name__icontains=advisor_name) |
                Q(advisor_set__coadvisor__name__icontains=advisor_name)
            ).distinct()
        return queryset

class StudentImportViewSet(ListModelMixin, CreateModelMixin, GenericViewSet):
    serializer_class = StudentTableSerializer
    parser_classes = [MultiPartParser]

    def get_queryset(self):
        if self.action == 'list':
            return Student.objects.none()
        else:
            return Student.objects.all()
        
    def list(self, request, *args, **kwargs):
        file_path = os.path.join(settings.BASE_DIR, 'templates', 'StudentTemplate.xlsx')
        wb = load_workbook(file_path)
        ws = wb.active

        if not "InstructorList" in wb.sheetnames:
            names_sheet = wb.create_sheet(title="InstructorList")
        else:
            names_sheet = wb["InstructorList"]

        names_sheet['C1'].value = "DO NOT EDIT THIS SHEET"
        names_sheet.sheet_state = 'hidden'

        instructors = list(Instructor.objects.order_by('name').values_list('name', 'emailId'))

        for index, instructor in enumerate(instructors, start=1):
            names_sheet[f'A{index}'].value = instructor[0]
            names_sheet[f'B{index}'].value = instructor[1]

        # defn = DefinedName("InstructorNames", attr_text=f'InstructorList!$A$1:$A${len(instructors)}')
        # wb.defined_names["InstructorNames"] = defn

        # dv_range = f'InstructorList!$A$1:$A${len(instructors)}'

        # dv = DataValidation(type="list", formula1=f'={dv_range}', allow_blank=True, showErrorMessage=True)
        # dv = DataValidation(type="list", formula1="InstructorNames", allow_blank=True, showErrorMessage=True)

        # dv.errorTitle = 'Invalid Option'
        # dv.error = 'This is not a valid option'

        # cell_range = CellRange(min_col=9, min_row=2, max_col=11, max_row=1048576)
        # dv.add(cell_range)

        # ws.add_data_validation(dv)

        wb.save(file_path)

        with open(file_path, 'rb') as file:
            response = HttpResponse(file.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="StudentTemplate.xlsx"'
            return response

    def camel_case(self, header):
        return header[0].lower() + header.replace('*','').replace(' ','')[1:]

    def get_advisor_email(self, name, instructor_list):
        for row in instructor_list.iter_rows(min_row=1, values_only=True):
            if name == row[0]:
                return row[1]
        return None

    def create(self, request, *args,**kwargs):
        if 'file' not in request.data:
            return Response({'error': 'No file uploaded'}, status=status.HTTP_400_BAD_REQUEST)
        
        uploaded_file = request.data['file']

        if not uploaded_file.name.endswith('.xlsx'):
            return Response({'error': 'Invalid file format. Please upload an Excel file (.xlsx)'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            workbook = load_workbook(uploaded_file)
            sheet = workbook.active

            # instructorList = workbook["InstructorList"]

            if "InstructorList" in workbook.sheetnames:
                instructorList = workbook["InstructorList"]
            else:
                raise Exception("Invalid template file uploaded. Download the template file and try again.")

            sheet_headers = [cell.value for cell in sheet[1]]
            required_columns = [
                'Roll Number*',
                'Name*',
                'Email Id*',
                'Gender*', 
                'Department*',
                'Joining Date*',
                'Batch Month*',
                'Batch Year*',
                'Advisor 1',
                'Advisor 2',
                'Coadvisor',
                'Educational Qualification',
                'Region',
                'Admission Through*',
                'Funding Type*',
                'Source Of Funding',
                'Contingency Points*',
                'Student Status*',
                'Thesis Submission Date',
                'Thesis Defence Date',
                'Year Of Leaving',
                'Comment',
                'Stipend Months',
                'Contingency Years'
            ]
            
            if not all(col in sheet_headers for col in required_columns):
                return Response({'error': 'Missing required columns in the Excel file'}, status=status.HTTP_400_BAD_REQUEST)

            invalid_rows = []
            processed_headers = [self.camel_case(header) for header in sheet_headers]

            for row in sheet.iter_rows(min_row=2, values_only=True):
                if all(cell is None for cell in row):
                    continue

                row_data = dict(zip(processed_headers, row))

                missing_headers = [header for header in required_columns if '*' in header and not row_data.get(self.camel_case(header))]
                if missing_headers:
                    invalid_rows.append({
                        'student': row_data,
                        'error': [f'Missing required data: {", ".join(missing_headers)}']
                    })
                    continue

                # if not all(row_data.get(self.camel_case(header)) for header in required_columns if '*' in header):
                #     invalid_rows.append({'student': row_data, 'error': ['Missing required data']})
                #     continue

                row_data['batch'] = f"{row_data.pop('batchMonth')} {row_data.pop('batchYear')}"
            
                for advisor_key in ['advisor1', 'advisor2', 'coadvisor']:
                    name = row_data.get(advisor_key)
                    if name:
                        email = self.get_advisor_email(name, instructorList)
                        if email:
                            row_data[f'{advisor_key}_emailId'] = email
                        else:
                            invalid_rows.append({'student': row_data, 'error': [f'{advisor_key}: {name} not found in Faculty List']})
                            continue
                    else:
                        row_data[f'{advisor_key}_emailId'] = 'null@iiitd.ac.in'

                for date_key in ['joiningDate', 'thesisSubmissionDate', 'thesisDefenceDate']:
                    if row_data[date_key] is not None:
                        row_data[date_key] = row_data[date_key].strftime('%d-%m-%Y')

                serializer = StudentTableSerializer(data=row_data)
                if serializer.is_valid():
                    serializer.save()
                else:
                    invalid_rows.append({'student': row_data, 'error': serializer.errors})

            if not invalid_rows:
                return Response({'message': 'Students successfully added'}, status=status.HTTP_201_CREATED)
            else:
                return Response({'error': 'Failed adding some students', 'invalid_data': invalid_rows}, status=status.HTTP_400_BAD_REQUEST)

        except Exception as e:
            return Response({'error': f'Error processing the Excel file: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class StudentExportViewSet(ListModelMixin, GenericViewSet):
    queryset = Student.objects.all()
    serializer_class = StudentTableSerializer
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['name', 'emailId', 'rollNumber', 'studentStatus', 'gender', 'department', 'batch', 'admissionThrough', 'region', 'fundingType', 'yearOfLeaving']
    search_fields = ['$name', '$emailId', '$rollNumber', '$advisor_set__advisor1__name', '$advisor_set__advisor2__name', '$advisor_set__coadvisor__name']

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        serializer = self.get_serializer(queryset, many=True)

        wb = Workbook()
        ws = wb.active
        ws.title = "Students"

        headers = ['Roll Number', 'Name', 'Email ID', 'Gender', 'Department', 'Advisor 1', 'Advisor 2', 'Coadvisor', 'Joining Date', 'Batch', 'Educational Qualification', 'Region', 'Admission Through', 'Funding Type', 'Source of Funding', 'Contingency Points', 'Student Status', 'Thesis Submission Date', 'Thesis Defence Date', 'Year of Leaving', 'Comment', 'Stipend Months', 'Contingency Years']
        ws.append(headers)

        for student in serializer.data:
            data = [
                student.get('rollNumber', ''),
                student.get('name', ''),
                student.get('emailId', ''),
                student.get('gender', ''),
                student.get('department', ''),
                student.get('advisor1', ''),
                student.get('advisor2', ''),
                student.get('coadvisor', ''),
                student.get('joiningDate', ''),
                student.get('batch', ''),
                student.get('educationalQualification', ''),
                student.get('region', ''),
                student.get('admissionThrough', ''),
                student.get('fundingType', ''),
                student.get('sourceOfFunding', ''),
                student.get('contingencyPoints', ''),
                student.get('studentStatus', ''),
                student.get('thesisSubmissionDate', ''),
                student.get('thesisDefenceDate', ''),
                student.get('yearOfLeaving', ''),
                student.get('comment', ''),
                student.get('stipendMonths', ''),
                student.get('contingencyYears', '')
            ]
            ws.append(data)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=students_data.xlsx'
        wb.save(response)

        return response

class InstructorViewSet(ModelViewSet):
    queryset = Instructor.objects.all()
    serializer_class = InstructorSerializer
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['emailId','name', 'department']
    search_fields = ['$emailId', '$name']

class AllInstructorViewSet(ReadOnlyModelViewSet):
    queryset = Instructor.objects.all()
    serializer_class = InstructorSerializer
    pagination_class = NoPagination
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['emailId','name', 'department']
    search_fields = ['$emailId', '$name']

class AdvisorViewSet(ModelViewSet):
    queryset = Advisor.objects.all()
    serializer_class = AdvisorSerializer
    lookup_field = 'student__rollNumber'
    lookup_url_kwarg = 'student__rollNumber'
    filter_backends = [DjangoFilterBackend, SearchFilter]
    filterset_fields = ['student', 'advisor1', 'advisor2', 'coadvisor']
    search_fields = ['$student__rollNumber']

class ComprehensiveViewSet(ModelViewSet):
    queryset = Comprehensive.objects.all()
    serializer_class = ComprehensiveSerializer
    pagination_class = NoPagination
    filter_backends = [SearchFilter]
    search_fields = ['$student__rollNumber']

class YearlyReviewViewSet(ModelViewSet):
    queryset = YearlyReview.objects.all()
    serializer_class = YearlyReviewSerializer
    pagination_class = NoPagination
    filter_backends = [SearchFilter, DjangoFilterBackend, OrderingFilter]
    filter_fields = ['reviewYear']
    search_fields = ['$student__rollNumber']

class ClassroomViewSet(ModelViewSet):
    queryset = Classroom.objects.all()
    serializer_class = ClassroomSerializer
    # pagination_class = NoPagination
    filter_backends = [DjangoFilterBackend, SearchFilter]
    filterset_fields = ['building', 'roomNo']
    search_fields = ['$building', '$roomNo']

    def create(self, request, *args, **kwargs):

        # Check if roomNo is alphanumeric
        roomNo = request.data['roomNo']
        capacity = request.data['capacity']
        if not roomNo.isalnum():
            return Response({'error': 'Incorrect Room No. Format: No Special Characters or Spaces Allowed. Valid Format Example: C101'}, status=status.HTTP_400_BAD_REQUEST)
        if not capacity.isdigit():
            return Response({'error': 'Incorrect Capacity Format: Capacity should be a Number.'}, status=status.HTTP_400_BAD_REQUEST)
        
        # Convert roomNo to uppercase
        request.data['roomNo'] = roomNo.upper()
        return super().create(request, *args, **kwargs)
    
    def update(self, request, *args, **kwargs):
        if "roomNo" in request.data.keys():
            roomNo = request.data['roomNo']
            if not roomNo.isalnum():
                return Response({'error': 'Incorrect Room No. Format: No Special Characters or Spaces Allowed. Valid Format Example: C101'}, status=status.HTTP_400_BAD_REQUEST)
            request.data['roomNo'] = roomNo.upper()

        if "capacity" in request.data.keys():
            capacity = request.data['capacity']
            if not capacity.isdigit():
                return Response({'error': 'Incorrect Capacity Format: Capacity should be a Number.'}, status=status.HTTP_400_BAD_REQUEST)
            
        return super().update(request, *args, **kwargs)


class StipendViewSet(ModelViewSet):
    queryset = Stipend.objects.all()
    serializer_class = StipendSerializer
    filter_backends = [DjangoFilterBackend, SearchFilter]
    filterset_fields = ['student__name', 'student__department', 'hostler', 'student__rollNumber', 'month', 'year']
    search_fields = ['$student__name', '$student__rollNumber']
    
    def list(self, request, *args, **kwargs):
        queryset = self.get_queryset()
        serializer = self.get_serializer(queryset, many=True)

        # Customizing the response
        response_data = []
        for stipend in serializer.data:
            student_id = stipend['student']
            student = Student.objects.get(pk=student_id)
            
            student_details = {
                'name': student.name,
                'rollNumber': student.rollNumber,
                'department': student.department,
            }
            stipend.update(student_details)
            response_data.append(stipend)

        return Response(response_data)

    def create(self, request, *args, **kwargs):
        # Assuming the request data is a list of Stipend objects
        stipend_data = request.data
        successful_entries = []
        failed_entries = []

        for stipend in stipend_data:
            student_id = stipend.get('student')
            month = stipend.get('month')
            year = stipend.get('year')

            # Check if a stipend entry already exists for the given student, month, and year
            if Stipend.objects.filter(student_id=student_id, month=month, year=year).exists():
                failed_entries.append({
                    'studentEntry': stipend,
                    'reason': f"Stipend Entry for month {month} and year {year} already exists for this student."
                })

                continue

            # Perform validations for each attribute
            serializer = self.get_serializer(data=stipend)
            if serializer.is_valid():
                # Months to be reduced.
                serializer.save()
                successful_entries.append(serializer.data)

                # Retrieve the Student instance and update its stipendMonths attribute
                student = Student.objects.get(pk=student_id)
                student.stipendMonths -= 1
                student.save()
            else:
                failed_entries.append({
                    'studentRollNumber': stipend.get('student__rollNumber'),
                    'reason': serializer.errors
                })

        response_data = {
            'successful_entries': successful_entries,
            'failed_entries': failed_entries
        }
        return Response(response_data, status=status.HTTP_201_CREATED)
    
    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()

        # Subtract amount from student.contingencyPoints
        student = instance.student
        student.stipendMonths += 1
        student.save()

        self.perform_destroy(instance)
        return Response(status=status.HTTP_204_NO_CONTENT)


class ContingencyViewSet(ModelViewSet):
    queryset = Contingency.objects.all()
    serializer_class = ContingencySerializer
    filter_backends = [DjangoFilterBackend, SearchFilter]
    filterset_fields = ['student__name', 'student__department', 'student__rollNumber', 'year']
    search_fields = ['$student__rollNumber', '$student__name']

    def list(self, request, *args, **kwargs):
        queryset = self.get_queryset()
        serializer = self.get_serializer(queryset, many=True)

        # Customizing the response
        response_data = []
        for contingency in serializer.data:
            student_id = contingency['student']
            # print(student_id)
            student = Student.objects.get(pk=student_id)
           
            
            student_details = {
                'name': student.name,
                'rollNumber': student.rollNumber,
                'department': student.department,
            }
            contingency.update(student_details)
            response_data.append(contingency)

        return Response(response_data)

    def create(self, request, *args, **kwargs):
        # Assuming the request data is a list of contingency objects
        contingency_data = request.data
        successful_entries = []
        failed_entries = []

        for contingency in contingency_data:
            student_id = contingency.get('student')
            year = contingency.get('year')

            # Check if a contingency entry already exists for the given student, and year
            if Contingency.objects.filter(student_id=student_id, year=year).exists():
                failed_entries.append({
                    'studentEntry': contingency,
                    'reason': f"Contingency Entry for year {year} already exists for this student."
                })

                continue

            # Perform validations for each attribute
            serializer = self.get_serializer(data=contingency)
            if serializer.is_valid():
                # Years to be reduced.
                serializer.save()
                successful_entries.append(serializer.data)

                # Retrieve the Student instance and update its contingencyPoints attribute
                student = Student.objects.get(pk=student_id)
                student.contingencyYears -= 1
                student.contingencyPoints += contingency.get('amount')
                student.save()
            else:
                failed_entries.append({
                    'studentRollNumber': contingency.get('student__rollNumber'),
                    'reason': serializer.errors
                })

        response_data = {
            'successful_entries': successful_entries,
            'failed_entries': failed_entries
        }
        return Response(response_data, status=status.HTTP_201_CREATED)
    
    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        original_amount = instance.amount

        # Don't allow editing student field
        request.data.pop('student')

        # Update other fields
        serializer = self.get_serializer(instance, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)

        new_amount = serializer.validated_data.get('amount', original_amount)
        difference = new_amount - original_amount

        # Update contingencyPoints in Student model
        student = instance.student
        student.contingencyPoints += difference
        student.save()

        return Response(serializer.data)

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()

        # Subtract amount from student.contingencyPoints
        student = instance.student
        student.contingencyPoints -= instance.amount
        student.contingencyYears += 1
        student.save()

        self.perform_destroy(instance)
        return Response(status=status.HTTP_204_NO_CONTENT)



class ContingencyLogsViewSet(ModelViewSet):
    queryset = ContingencyLogs.objects.all()
    serializer_class = ContingencyLogsSerializer
    pagination_class = NoPagination
    filter_backends = [SearchFilter]
    search_fields = ['$student__rollNumber']

    def perform_destroy(self, instance):
        if instance.santionedAmount is not None:
            student = instance.student
            student.contingencyPoints += instance.santionedAmount
            student.save()
        instance.delete()


class EligibleStudentStipendViewSet(ReadOnlyModelViewSet):
    serializer_class = StudentSerializer

    def get_queryset(self):
        # Existing filtering logic
        queryset = Student.objects.filter(studentStatus="Active", stipendMonths__gt=0, fundingType="Institute")
        
        # Extract month and year from request
        month = self.request.query_params.get('month', None)
        year = self.request.query_params.get('year', None)

        # If both month and year are provided, filter students without stipend for that month and year
        if month is not None and year is not None:
            month = int(month)
            year = int(year)
            # Exclude students who have a stipend record for the given month and year
            queryset = queryset.exclude(stipend__month=month, stipend__year=year)

        return queryset
    
    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        month = self.request.query_params.get('month', None)
        year = self.request.query_params.get('year', None)


        response_data = []
        for student in queryset:
            student_data = {
                'id': student.id,
                'name': student.name,
                'rollNumber': student.rollNumber,
                'month': month,
                'year': year,
                'joiningDate': student.joiningDate,
                'department': student.department,
                'hra': 0,
                'hostler': 'YES',
                'comment':""
            }

            try:
                comprehensive_review = student.comprehensive_review
                student_data['comprehensiveExamDate'] = comprehensive_review.dateOfReview
                student_data['baseAmount'] = 42000
            except Comprehensive.DoesNotExist:
                student_data['comprehensiveExamDate'] = None
                student_data['baseAmount'] = 37000

            response_data.append(student_data)

        return Response(response_data, status=status.HTTP_200_OK)
    


class EligibleStudentContingencyViewSet(ReadOnlyModelViewSet):
    serializer_class = StudentSerializer

    def get_queryset(self):
        # Existing filtering logic
        queryset = Student.objects.filter(studentStatus="Active", contingencyYears__gt=0)
        
        # Extract year from request
        year = self.request.query_params.get('year', None)

        # If both year are provided, filter students without stipend for that year
        if year is not None:
            year = int(year)
            # Exclude students who have a stipend record for the given month and year
            queryset = queryset.exclude(contingency__year=year)

        return queryset
    
    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        year = self.request.query_params.get('year', None)


        response_data = []
        for student in queryset:
            student_data = {
                'id': student.id,
                'name': student.name,
                'rollNumber': student.rollNumber,
                'year': year,
                'joiningDate': student.joiningDate,
                'department': student.department,
                'amount':20000,
                'comment':""
            }

            response_data.append(student_data)

        return Response(response_data, status=status.HTTP_200_OK)
     
class AllotmentViewSet(ViewSet):
    parser_classes = [MultiPartParser]

    def create(self, request, format=None): # In the post method of the APIView, the format=None parameter indicates the format in which the response should be returned.
        TARatio = request.data.get('TARatio')

        # Define the directory to save the files
        invigilation_files_dir = os.path.join(settings.MEDIA_ROOT, 'invigilationFiles')
        os.makedirs(invigilation_files_dir, exist_ok=True)

        # Define the filenames to save the uploaded files
        file_names = {
            'file1': invigilation_files_dir + '\Classroom.xlsx',
        }

        for key, file in request.FILES.items():
            file_name = file_names.get(key)
            if file_name:
                with open(file_name, 'wb+') as destination:
                    for chunk in file.chunks():
                        destination.write(chunk)

        # Run the python script with TARatio as an argument
        script_path = os.path.join(settings.BASE_DIR, 'algorithm', 'allocationAlgorithm.py')
        subprocess.call(['python', script_path, '--TARatio', TARatio])

        # Define the path to the generated file
        generated_file_path = os.path.join(invigilation_files_dir, 'InvigilatorList.xlsx')

        # Check if the generated file exists before attempting to open it
        if os.path.exists(generated_file_path):
            with open(generated_file_path, 'rb') as f:
                response = HttpResponse(f.read(), content_type='application/octet-stream')
                response['Content-Disposition'] = f'attachment; filename={os.path.basename(generated_file_path)}'

            # Delete the files after sending the response
            for file_name in file_names.values():
                if os.path.exists(file_name):
                    os.remove(file_name)
            if os.path.exists(generated_file_path):
                os.remove(generated_file_path)

            return response
        else:
            return Response({'message': 'Generated file not found'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)